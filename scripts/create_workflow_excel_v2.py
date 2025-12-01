# -*- coding: utf-8 -*-
"""
TwinCAT PLC 워크플로우 문서를 스타일이 적용된 엑셀 파일로 생성
전체 흐름도 시트 포함 버전
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# 색상 정의
COLORS = {
    'primary': '2E75B6',      # 파란색 (헤더)
    'secondary': '5B9BD5',    # 연한 파란색
    'accent': 'ED7D31',       # 주황색 (강조)
    'success': '70AD47',      # 녹색 (완료/성공)
    'warning': 'FFC000',      # 노란색 (주의)
    'danger': 'C00000',       # 빨간색 (위험)
    'light_gray': 'F2F2F2',   # 연한 회색 (배경)
    'dark_gray': '404040',    # 진한 회색 (텍스트)
    'white': 'FFFFFF',
    'light_blue': 'DEEBF7',   # 연한 파란 배경
    'light_green': 'E2EFDA',  # 연한 녹색 배경
    'light_orange': 'FCE4D6', # 연한 주황 배경
    'scr': '4472C4',          # SCR 파란색
    'wise': 'ED7D31',         # WISE 주황색
    'develop': '70AD47',      # 개발 녹색
    'test': '5B9BD5',         # 테스트 연파란
    'end_success': '00B050',  # 완료 녹색
    'end_fail': 'C00000',     # 종료 빨간색
    'arrow': '7F7F7F',        # 화살표 회색
}

# 테두리 스타일
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

medium_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

def create_header_style():
    """헤더 스타일 생성"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=12),
        'fill': PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border
    }

def create_subheader_style():
    """서브헤더 스타일"""
    return {
        'font': Font(bold=True, color=COLORS['dark_gray'], size=11),
        'fill': PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border
    }

def create_title_style():
    """제목 스타일"""
    return {
        'font': Font(bold=True, color=COLORS['primary'], size=16),
        'alignment': Alignment(horizontal='center', vertical='center'),
    }

def create_section_title_style():
    """섹션 제목 스타일"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=13),
        'fill': PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': thin_border
    }

def apply_style(cell, style_dict):
    """셀에 스타일 적용"""
    for key, value in style_dict.items():
        setattr(cell, key, value)

def set_column_widths(ws, widths):
    """컬럼 너비 설정"""
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def create_flowchart_sheet(wb):
    """전체 워크플로우 흐름도 시트 생성"""
    ws = wb.create_sheet("전체 흐름도", 0)

    # 컬럼 너비 설정 (A~T = 20개 컬럼)
    for col in range(1, 21):
        ws.column_dimensions[get_column_letter(col)].width = 4

    # 행 높이 설정
    for row in range(1, 70):
        ws.row_dimensions[row].height = 18

    # 제목
    ws.merge_cells('A1:T1')
    ws['A1'] = "TwinCAT PLC 개발 워크플로우 전체 흐름도"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 범례
    ws.merge_cells('A3:C3')
    ws['A3'] = "범례:"
    ws['A3'].font = Font(bold=True, size=11)

    # 범례 항목
    legend_items = [
        ('D3:E3', '시작/SCR', COLORS['scr']),
        ('F3:G3', 'WISE/승인', COLORS['wise']),
        ('H3:I3', '개발', COLORS['develop']),
        ('J3:K3', '테스트', COLORS['test']),
        ('L3:M3', '완료', COLORS['end_success']),
        ('N3:O3', '종료', COLORS['end_fail']),
    ]

    for cell_range, text, color in legend_items:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = text
        cell.fill = PatternFill(start_color=color, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 흐름도 그리기 함수들
    def draw_box(row, col_start, col_end, text, color, is_terminal=False):
        """상자 그리기"""
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start)
        cell.value = text
        cell.fill = PatternFill(start_color=color, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = medium_border if is_terminal else thin_border

    def draw_arrow_down(row, col, symbol='│'):
        """아래 화살표"""
        cell = ws.cell(row=row, column=col)
        cell.value = symbol
        cell.font = Font(color=COLORS['arrow'], size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def draw_arrow_symbol(row, col, symbol):
        """화살표 기호"""
        cell = ws.cell(row=row, column=col)
        cell.value = symbol
        cell.font = Font(color=COLORS['arrow'], size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ===== 흐름도 시작 =====
    base_row = 5
    center_col = 10  # 중앙 컬럼

    # 시작
    draw_box(base_row, center_col-1, center_col+1, "시작", COLORS['scr'])
    draw_arrow_down(base_row+1, center_col, '▼')

    # SCR
    draw_box(base_row+2, center_col-1, center_col+1, "SCR", COLORS['scr'])
    draw_arrow_down(base_row+3, center_col, '▼')

    # WISE (분기점)
    draw_box(base_row+4, center_col-1, center_col+1, "WISE", COLORS['wise'])

    # WISE에서 왼쪽으로 -> 기각
    draw_arrow_symbol(base_row+4, center_col-2, '◀──')
    draw_box(base_row+4, 3, 5, "기각 (종료)", COLORS['end_fail'], True)

    # WISE에서 오른쪽으로 -> LEADER APPROVE
    draw_arrow_symbol(base_row+4, center_col+2, '──▶')
    draw_box(base_row+4, 15, 18, "LEADER APPROVE", COLORS['wise'])

    # WISE에서 아래로
    draw_arrow_down(base_row+5, center_col, '│')

    # LEADER APPROVE 분기
    draw_arrow_down(base_row+5, 17, '│')
    draw_arrow_symbol(base_row+6, 16, '┌──┴──┐')

    # LEADER APPROVE -> 기각
    draw_box(base_row+7, 14, 15, "기각", COLORS['end_fail'], True)

    # LEADER APPROVE -> 승인 (WISE로 합류)
    draw_box(base_row+7, 17, 18, "승인", COLORS['end_success'])
    draw_arrow_down(base_row+8, 17, '│')
    draw_arrow_symbol(base_row+9, 13, '◀────────┘')

    # SW DEVELOP
    draw_arrow_down(base_row+6, center_col, '▼')
    draw_box(base_row+10, center_col-2, center_col+2, "SW DEVELOP", COLORS['develop'])

    # SW DEVELOP -> SCR 미적용
    draw_arrow_symbol(base_row+10, center_col-3, '◀──')
    draw_box(base_row+10, 2, 5, "SCR 미적용 (종료)", COLORS['end_fail'], True)

    # SW DEVELOP -> MERGED
    draw_arrow_down(base_row+11, center_col, '▼')
    draw_box(base_row+12, center_col-1, center_col+1, "MERGED", COLORS['develop'])

    # MERGED -> SCR 미적용
    draw_arrow_symbol(base_row+12, center_col-2, '◀──')
    draw_box(base_row+12, 2, 5, "SCR 미적용 (종료)", COLORS['end_fail'], True)

    # MERGED 분기 (3방향)
    draw_arrow_down(base_row+13, center_col, '│')
    draw_arrow_symbol(base_row+14, center_col-3, '┌────┼────┐')

    # MERGED -> 설비 검증 (왼쪽)
    draw_arrow_down(base_row+14, center_col-3, '▼')
    draw_box(base_row+15, center_col-5, center_col-2, "설비 검증", COLORS['test'])

    # MERGED -> SW RESOLVED (오른쪽)
    draw_arrow_down(base_row+14, center_col+3, '▼')
    draw_box(base_row+15, center_col+1, center_col+4, "SW RESOLVED", COLORS['develop'])

    # 설비 검증 -> SW RESOLVED
    draw_arrow_down(base_row+16, center_col-3, '▼')
    draw_box(base_row+17, center_col-5, center_col-2, "SW RESOLVED", COLORS['develop'])

    # SW RESOLVED 분기 (오른쪽)
    draw_arrow_symbol(base_row+15, center_col+5, '──┐')
    draw_arrow_down(base_row+16, center_col+6, '│')
    draw_arrow_symbol(base_row+17, center_col+5, '┌─┴─┐')

    # SW RESOLVED -> SCR 미적용
    draw_box(base_row+18, center_col+3, center_col+5, "SCR 미적용", COLORS['end_fail'], True)

    # SW RESOLVED -> SW RESOLVED 승인 -> SCR 완료
    draw_box(base_row+18, center_col+6, center_col+9, "SW RESOLVED 승인", COLORS['end_success'])
    draw_arrow_down(base_row+19, center_col+7, '▼')
    draw_box(base_row+20, center_col+6, center_col+9, "SCR 완료 (종료)", COLORS['end_success'], True)

    # 설비 검증 후 SW RESOLVED 분기
    draw_arrow_symbol(base_row+17, center_col-1, '──┐')
    draw_arrow_down(base_row+18, center_col, '│')
    draw_arrow_symbol(base_row+19, center_col-1, '┌─┴─┐')

    # 왼쪽 SW RESOLVED -> SCR 미적용
    draw_box(base_row+20, center_col-4, center_col-2, "SCR 미적용", COLORS['end_fail'], True)

    # 왼쪽 SW RESOLVED -> SW RESOLVED 승인 -> SCR 완료
    draw_box(base_row+20, center_col, center_col+2, "승인", COLORS['end_success'])
    draw_arrow_down(base_row+21, center_col+1, '▼')
    draw_box(base_row+22, center_col-1, center_col+3, "SCR 완료 (종료)", COLORS['end_success'], True)

    # SQA TEST 경로
    draw_arrow_symbol(base_row+23, center_col, '─────┬─────')
    draw_arrow_down(base_row+24, center_col, '▼')
    draw_box(base_row+25, center_col-1, center_col+1, "SQA TEST", COLORS['test'])

    # SQA TEST 분기
    draw_arrow_down(base_row+26, center_col, '│')
    draw_arrow_symbol(base_row+27, center_col-2, '┌──┴──┐')

    # TEST PASS
    draw_arrow_down(base_row+27, center_col-2, '▼')
    draw_box(base_row+28, center_col-4, center_col-1, "TEST PASS", COLORS['end_success'])

    # TEST 결과 검토
    draw_arrow_down(base_row+27, center_col+2, '▼')
    draw_box(base_row+28, center_col+1, center_col+4, "TEST 결과 검토", COLORS['test'])

    # TEST 결과 검토 분기
    draw_arrow_down(base_row+29, center_col+2, '│')
    draw_arrow_symbol(base_row+30, center_col+1, '┌──┴──┐')

    # NG시 -> SW DEVELOP 복귀
    draw_arrow_down(base_row+30, center_col+1, '▼')
    draw_box(base_row+31, center_col, center_col+2, "NG시", COLORS['end_fail'])
    ws.merge_cells(start_row=base_row+32, start_column=center_col+1, end_row=base_row+32, end_column=center_col+3)
    cell = ws.cell(row=base_row+32, column=center_col+1)
    cell.value = "↑ SW DEVELOP 복귀"
    cell.font = Font(italic=True, size=9, color=COLORS['arrow'])
    cell.alignment = Alignment(horizontal='center')

    # 재테스트 -> SQA TEST 복귀
    draw_arrow_down(base_row+30, center_col+4, '▼')
    draw_box(base_row+31, center_col+3, center_col+5, "재테스트", COLORS['test'])
    ws.merge_cells(start_row=base_row+32, start_column=center_col+4, end_row=base_row+32, end_column=center_col+6)
    cell = ws.cell(row=base_row+32, column=center_col+4)
    cell.value = "↑ SQA TEST 복귀"
    cell.font = Font(italic=True, size=9, color=COLORS['arrow'])
    cell.alignment = Alignment(horizontal='center')

    # TEST PASS -> SCR 완료
    draw_arrow_down(base_row+29, center_col-2, '▼')
    draw_box(base_row+30, center_col-4, center_col-1, "SCR 완료 (종료)", COLORS['end_success'], True)

    # 주요 경로 설명 추가
    desc_row = base_row + 35
    ws.merge_cells(f'A{desc_row}:T{desc_row}')
    ws[f'A{desc_row}'] = "주요 경로 설명"
    apply_style(ws[f'A{desc_row}'], create_section_title_style())

    paths = [
        ['기본 경로', 'SCR → WISE → SW DEVELOP → MERGED → SW RESOLVED → SW RESOLVED 승인 → SCR 완료'],
        ['Level 1 (사내검증 생략)', 'SW DEVELOP → MERGED → SW RESOLVED (사내검증 없이)'],
        ['Level 2/3 (사내검증 필수)', 'SW DEVELOP → MERGED → 사내검증(설비/SQA) → SW RESOLVED'],
        ['설비 검증 경로', 'MERGED → 설비 검증 → SW RESOLVED → SQA TEST/SW RESOLVED 승인'],
        ['SQA TEST 경로', 'MERGED/SW RESOLVED → SQA TEST → TEST PASS → SCR 완료'],
        ['TEST NG 경로', 'TEST 결과 검토 → NG시 → SW DEVELOP (재개발)'],
        ['재테스트 경로', 'TEST 결과 검토 → 재테스트 → SQA TEST'],
    ]

    headers = ['경로명', '흐름']
    for col_idx, header in enumerate(headers, 1):
        if col_idx == 1:
            ws.merge_cells(start_row=desc_row+1, start_column=1, end_row=desc_row+1, end_column=4)
            cell = ws.cell(row=desc_row+1, column=1)
        else:
            ws.merge_cells(start_row=desc_row+1, start_column=5, end_row=desc_row+1, end_column=20)
            cell = ws.cell(row=desc_row+1, column=5)
        cell.value = header
        apply_style(cell, create_header_style())

    for row_idx, path in enumerate(paths, desc_row+2):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        cell1 = ws.cell(row=row_idx, column=1)
        cell1.value = path[0]
        cell1.border = thin_border
        cell1.font = Font(bold=True)
        cell1.fill = PatternFill(start_color=COLORS['light_blue'], fill_type='solid')

        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=20)
        cell2 = ws.cell(row=row_idx, column=5)
        cell2.value = path[1]
        cell2.border = thin_border

    return ws

def create_workflow_sheet(wb):
    """워크플로우 개요 시트 생성"""
    ws = wb.create_sheet("워크플로우 개요")

    # 컬럼 너비 설정
    set_column_widths(ws, [5, 20, 25, 30, 25, 20])

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "TwinCAT PLC 개발 워크플로우 개선안"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # 문서 정보
    ws.merge_cells('A3:F3')
    ws['A3'] = "문서 정보"
    apply_style(ws['A3'], create_section_title_style())

    doc_info = [
        ['버전', '1.0', '작성일', '2025-11-27', '', ''],
        ['대상 독자', '개발자, 관리자, 의사결정권자 (100명+)', '목적', '현실적이고 점진적인 워크플로우 개선', '', ''],
    ]

    for row_idx, row_data in enumerate(doc_info, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in [1, 3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=COLORS['light_gray'], fill_type='solid')

    # 기본 흐름
    ws.merge_cells('A7:F7')
    ws['A7'] = "기본 워크플로우 흐름"
    apply_style(ws['A7'], create_section_title_style())

    flow_header = ['단계', '상태', '설명', '다음 가능 단계', '담당', '소요시간']
    for col_idx, header in enumerate(flow_header, 1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    flow_data = [
        ['1', 'SCR', '소프트웨어 변경 요청', 'WISE', '요청자', '1일'],
        ['2', 'WISE', '기술 검토 위원회', 'SW DEVELOP, LEADER APPROVE, 기각', '위원회', '1-3일'],
        ['3', 'LEADER APPROVE', '리더 승인 (필요시)', 'SW DEVELOP, 기각', '리더', '1일'],
        ['4', 'SW DEVELOP', '소프트웨어 개발', 'MERGED, SCR 미적용', '개발자', '1-14일'],
        ['5', 'MERGED', '코드 병합 완료', '설비 검증, SQA TEST, SW RESOLVED', '개발자', '1일'],
        ['6', '설비 검증', '실제 설비에서 검증', 'SW RESOLVED', 'QA', '1-3일'],
        ['7', 'SW RESOLVED', '개발 완료 상태', 'SQA TEST, SW RESOLVED 승인', '개발자', '1일'],
        ['8', 'SQA TEST', '품질 보증 테스트', 'TEST PASS, TEST 결과 검토', 'SQA', '2-4일'],
        ['9', 'TEST PASS', '테스트 통과', 'SCR 완료', 'SQA', '-'],
        ['10', 'SCR 완료', '최종 완료 (배포)', '-', '관리자', '-'],
    ]

    status_colors = {
        'SCR': COLORS['light_blue'],
        'WISE': COLORS['light_orange'],
        'LEADER APPROVE': COLORS['light_orange'],
        'SW DEVELOP': COLORS['light_green'],
        'MERGED': COLORS['light_green'],
        '설비 검증': COLORS['light_blue'],
        'SW RESOLVED': COLORS['light_green'],
        'SQA TEST': COLORS['light_blue'],
        'TEST PASS': 'C6EFCE',
        'SCR 완료': 'C6EFCE',
    }

    for row_idx, row_data in enumerate(flow_data, 9):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            if col_idx == 2 and value in status_colors:
                cell.fill = PatternFill(start_color=status_colors[value], fill_type='solid')
                cell.font = Font(bold=True)

    ws.row_dimensions[8].height = 25
    for i in range(9, 19):
        ws.row_dimensions[i].height = 30

    return ws

def create_change_classification_sheet(wb):
    """변경 분류 체계 시트 생성"""
    ws = wb.create_sheet("변경 분류 체계")

    set_column_widths(ws, [5, 18, 35, 35, 25, 20])

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "변경 분류 체계 (Level 기반)"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # 설명
    ws.merge_cells('A3:F3')
    ws['A3'] = "현재 워크플로우에 이미 적용 중인 변경 분류 체계입니다."
    ws['A3'].font = Font(italic=True, size=11)

    # 분류 체계 테이블
    ws.merge_cells('A5:F5')
    ws['A5'] = "Level 기반 분류 체계"
    apply_style(ws['A5'], create_section_title_style())

    headers = ['Level', '경로', '특징', '대상', '승인', '테스트']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    classification_data = [
        ['Level 1', 'DEVELOP → MERGED → SW RESOLVED', '사내검증 생략', '1-2개 파일, 기존 로직 수정', '팀 리더', '자체 테스트'],
        ['Level 2', 'DEVELOP → MERGED → 사내검증 → SW RESOLVED', '사내검증 필수', '2-10개 파일, 신규 기능', 'WISE', '설비 검증 선택'],
        ['Level 3', 'DEVELOP → MERGED → 사내검증 → SW RESOLVED', '전체 검증 필수', '아키텍처 변경, 다수 파일', 'WISE + LEADER', '설비 + SQA 필수'],
    ]

    level_colors = {
        'Level 1': 'C6EFCE',
        'Level 2': 'FFEB9C',
        'Level 3': 'FFC7CE',
    }

    for row_idx, row_data in enumerate(classification_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            if col_idx == 1:
                cell.fill = PatternFill(start_color=level_colors[value], fill_type='solid')
                cell.font = Font(bold=True)

    for i in range(6, 10):
        ws.row_dimensions[i].height = 35

    # 예시
    ws.merge_cells('A12:F12')
    ws['A12'] = "Level별 예시"
    apply_style(ws['A12'], create_section_title_style())

    example_headers = ['Level', '예시 1', '예시 2', '예시 3', '', '']
    for col_idx, header in enumerate(example_headers, 1):
        cell = ws.cell(row=13, column=col_idx, value=header)
        apply_style(cell, create_subheader_style())

    examples = [
        ['Level 1', '파라미터 값 변경', '타이머 조정', '단순 버그 수정', '', ''],
        ['Level 2', '새 Function Block 추가', '알람 로직 변경', '기존 기능 확장', '', ''],
        ['Level 3', '새 설비 제어 추가', '통신 프로토콜 변경', '아키텍처 재설계', '', ''],
    ]

    for row_idx, row_data in enumerate(examples, 14):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.fill = PatternFill(start_color=level_colors[row_data[0]], fill_type='solid')
                cell.font = Font(bold=True)

    return ws

def create_analysis_sheet(wb):
    """장단점 분석 시트 생성"""
    ws = wb.create_sheet("장단점 분석")

    set_column_widths(ws, [5, 25, 40, 15, 20])

    # 제목
    ws.merge_cells('A1:E1')
    ws['A1'] = "현재 워크플로우 장단점 분석"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # 장점 섹션
    ws.merge_cells('A3:E3')
    ws['A3'] = "✓ 장점 (Strengths)"
    apply_style(ws['A3'], create_section_title_style())
    ws['A3'].fill = PatternFill(start_color=COLORS['success'], fill_type='solid')

    strength_headers = ['#', '항목', '효과', '중요도', '비고']
    for col_idx, header in enumerate(strength_headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, create_header_style())
        cell.fill = PatternFill(start_color=COLORS['success'], fill_type='solid')

    strengths = [
        ['1', '다단계 품질 게이트', '프로덕션 버그 최소화, 결함 조기 발견', '★★★★★', '핵심 강점'],
        ['2', '명확한 승인 체계', '거버넌스 확립, 책임 소재 명확', '★★★★☆', '규제 준수'],
        ['3', '추적 가능성', '감사 대응, 변경 이력 완전 보존', '★★★★★', '필수 요구사항'],
        ['4', 'Level 기반 분류 체계', '리스크 수준별 검증 차등화', '★★★★☆', '효율성 확보'],
        ['5', '유연한 검증 경로', '하드웨어 의존/비의존 경로 분리', '★★★☆☆', '유연성'],
    ]

    for row_idx, row_data in enumerate(strengths, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=COLORS['light_green'], fill_type='solid')

    # 단점 섹션
    ws.merge_cells('A11:E11')
    ws['A11'] = "✗ 단점 (Weaknesses)"
    apply_style(ws['A11'], create_section_title_style())
    ws['A11'].fill = PatternFill(start_color=COLORS['danger'], fill_type='solid')

    for col_idx, header in enumerate(strength_headers, 1):
        cell = ws.cell(row=12, column=col_idx, value=header)
        apply_style(cell, create_header_style())
        cell.fill = PatternFill(start_color=COLORS['danger'], fill_type='solid')

    weaknesses = [
        ['1', 'WISE 직렬화', '모든 SCR이 회의 대기, 2-3일 지연', '★★★★★', '최우선 개선'],
        ['2', '긴급 대응 부재', 'Fast Track 프로세스 없음', '★★★★★', '즉시 개선 필요'],
        ['3', '순차 검증', '설비 검증 → SQA 순차 실행 (7일)', '★★★★☆', '병렬화 가능'],
        ['4', '수동 테스트 의존', '재현성 낮음, 반복 비용 발생', '★★★☆☆', '자동화 필요'],
        ['5', '설비 접근 제한', '야간/주말 작업 강제', '★★★★★', '인프라 투자'],
    ]

    for row_idx, row_data in enumerate(weaknesses, 13):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=COLORS['light_orange'], fill_type='solid')

    # 비용 분석
    ws.merge_cells('A20:E20')
    ws['A20'] = "연간 숨겨진 비용 분석"
    apply_style(ws['A20'], create_section_title_style())

    cost_headers = ['#', '비용 항목', '연간 추정 비용', '비율', '']
    for col_idx, header in enumerate(cost_headers, 1):
        cell = ws.cell(row=21, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    costs = [
        ['1', '긴급 대응 야근/주말 근무', '₩30,000,000', '27%', ''],
        ['2', '테스트 반복 인건비', '₩25,000,000', '23%', ''],
        ['3', '프로덕션 버그 대응', '₩40,000,000', '36%', ''],
        ['4', '커뮤니케이션 오버헤드', '₩15,000,000', '14%', ''],
        ['합계', '', '₩110,000,000', '100%', ''],
    ]

    for row_idx, row_data in enumerate(costs, 22):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_data[0] == '합계':
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=COLORS['warning'], fill_type='solid')

    return ws

def create_as_is_to_be_sheet(wb):
    """AS-IS / TO-BE 시트 생성"""
    ws = wb.create_sheet("AS-IS TO-BE")

    set_column_widths(ws, [5, 20, 35, 35, 20])

    # 제목
    ws.merge_cells('A1:E1')
    ws['A1'] = "AS-IS / TO-BE 개선안"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # 개선 항목들
    improvements = [
        {
            'title': '1. WISE 승인 프로세스',
            'as_is': '모든 SCR → WISE 회의 대기 (2-3일)\n주 1-2회 정기 회의에 종속\n긴급 대응 불가',
            'to_be': 'Level별 차등 승인:\n- Level 1: 팀 리더 비동기 승인\n- Level 2: WISE 비동기 검토\n- Level 3: WISE 정식 회의\n- Fast Track: 리더 구두 승인',
            'effect': '대기 시간 66% 감소'
        },
        {
            'title': '2. 검증 프로세스',
            'as_is': '설비 검증 (3일) → SQA TEST (4일)\n순차 실행으로 총 7일 소요',
            'to_be': '병렬 검증:\n설비 검증 (3일) ─┐\n                     ├→ 통합 (0.5일)\nSQA TEST (4일) ─┘\n총 4.5일',
            'effect': '검증 시간 36% 단축'
        },
        {
            'title': '3. 긴급 대응',
            'as_is': '공식 경로 없음\n비공식 수정 → 추적 불가\n감사 리스크 발생',
            'to_be': 'Fast Track 프로세스:\n1. 리더 구두 승인 (즉시)\n2. 개발 착수\n3. 48시간 내 SCR 작성\n4. 주간 회의 사후 보고',
            'effect': '대응 시간 98% 단축\n(7일 → 4시간)'
        },
        {
            'title': '4. 테스트 프로세스',
            'as_is': '수동 테스트 의존\n재현성 낮음\n평균 1.5회 반복',
            'to_be': '자동화 도입 (TcUnit):\n- 단위 테스트 자동 실행\n- 회귀 테스트 포함\n- 표준 시나리오 템플릿',
            'effect': '테스트 시간 50% 단축\n반복 1.5회 → 1.1회'
        },
        {
            'title': '5. 코드 리뷰',
            'as_is': '코드 리뷰 프로세스 미정의\n개발자 독립 작업 후 병합\n지식 공유 없음',
            'to_be': 'PR 기반 리뷰 도입:\n- Level 1: 자기 검토\n- Level 2: 동료 1인 리뷰\n- Level 3: 시니어 + 동료',
            'effect': '버그 조기 발견\n속인화 해소'
        },
    ]

    row = 3
    for item in improvements:
        # 섹션 제목
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = item['title']
        apply_style(ws[f'A{row}'], create_section_title_style())
        row += 1

        # 헤더
        headers = ['', 'AS-IS (현재)', 'TO-BE (개선)', '개선 효과', '']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            if header:
                apply_style(cell, create_subheader_style())
                if header == 'AS-IS (현재)':
                    cell.fill = PatternFill(start_color='FFC7CE', fill_type='solid')
                elif header == 'TO-BE (개선)':
                    cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')
        row += 1

        # 내용
        cell_as_is = ws.cell(row=row, column=2, value=item['as_is'])
        cell_as_is.border = thin_border
        cell_as_is.alignment = Alignment(vertical='top', wrap_text=True)
        cell_as_is.fill = PatternFill(start_color='FFF2CC', fill_type='solid')

        cell_to_be = ws.cell(row=row, column=3, value=item['to_be'])
        cell_to_be.border = thin_border
        cell_to_be.alignment = Alignment(vertical='top', wrap_text=True)
        cell_to_be.fill = PatternFill(start_color='E2EFDA', fill_type='solid')

        cell_effect = ws.cell(row=row, column=4, value=item['effect'])
        cell_effect.border = thin_border
        cell_effect.alignment = Alignment(vertical='top', wrap_text=True)
        cell_effect.font = Font(bold=True, color=COLORS['success'])

        ws.row_dimensions[row].height = 100
        row += 2

    return ws

def create_roadmap_sheet(wb):
    """실행 로드맵 시트 생성"""
    ws = wb.create_sheet("실행 로드맵")

    set_column_widths(ws, [5, 20, 40, 15, 15, 20])

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "3단계 점진적 개선 로드맵"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    phases = [
        {
            'name': 'Phase 0: 즉시 적용',
            'cost': '₩0',
            'duration': '1주',
            'color': COLORS['success'],
            'items': [
                ['Fast Track 프로세스 신설', '긴급 대응 98% 단축', '프로세스'],
                ['변경 분류 체계 명확화', '기존 Level 체계 강화', '프로세스'],
                ['WISE 비동기 검토 허용', '대기 시간 66% 감소', '프로세스'],
            ]
        },
        {
            'name': 'Phase 1: 프로세스 개선',
            'cost': '₩5,000,000',
            'duration': '1-2개월',
            'color': COLORS['secondary'],
            'items': [
                ['Git 브랜치 전략 표준화', '통합 충돌 감소', '도구'],
                ['코드 리뷰 프로세스 도입', '품질 향상, 지식 공유', '프로세스'],
                ['테스트 시나리오 표준화', '재현성 향상', '프로세스'],
                ['설비 검증 + SQA 병렬화', '검증 36% 단축', '프로세스'],
            ]
        },
        {
            'name': 'Phase 2: 기본 자동화',
            'cost': '₩5-15,000,000',
            'duration': '3-6개월',
            'color': COLORS['accent'],
            'items': [
                ['자동 빌드 시스템', '수동 빌드 제거', '인프라'],
                ['TcUnit 단위 테스트 도입', '테스트 50% 단축', '도구'],
                ['모니터링 대시보드', '메트릭 시각화', '인프라'],
                ['간소화 스테이징 환경', '배포 전 검증', '인프라'],
            ]
        },
        {
            'name': 'Phase 3: 고급 자동화 (선택)',
            'cost': '₩15-50,000,000',
            'duration': '6-12개월',
            'color': COLORS['primary'],
            'items': [
                ['완전 CI/CD 파이프라인', '배포 완전 자동화', '인프라'],
                ['점진적 배포 (PLC 특화)', '배포 리스크 감소', '프로세스'],
                ['자동 롤백 시스템', '장애 복구 자동화', '인프라'],
            ]
        },
    ]

    row = 3
    for phase in phases:
        # Phase 헤더
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"{phase['name']} | 비용: {phase['cost']} | 기간: {phase['duration']}"
        apply_style(ws[f'A{row}'], create_section_title_style())
        ws[f'A{row}'].fill = PatternFill(start_color=phase['color'], fill_type='solid')
        row += 1

        # 항목 헤더
        item_headers = ['#', '항목', '효과', '유형', '', '']
        for col_idx, header in enumerate(item_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            if header:
                apply_style(cell, create_subheader_style())
        row += 1

        # 항목들
        for idx, item in enumerate(phase['items'], 1):
            ws.cell(row=row, column=1, value=idx).border = thin_border
            for col_idx, value in enumerate(item, 2):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
            row += 1

        row += 1

    return ws

def create_roi_sheet(wb):
    """ROI 분석 시트 생성"""
    ws = wb.create_sheet("ROI 분석")

    set_column_widths(ws, [5, 25, 20, 20, 20, 15])

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "투자 대비 효과 분석 (ROI)"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # Phase별 ROI
    ws.merge_cells('A3:F3')
    ws['A3'] = "Phase별 투자 및 회수"
    apply_style(ws['A3'], create_section_title_style())

    headers = ['Phase', '투자 비용', '연간 절감', '회수 기간', 'ROI', '']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    roi_data = [
        ['Phase 0', '₩0', '₩30,000,000', '즉시', '∞', ''],
        ['Phase 1', '₩5,000,000', '₩20,000,000 추가', '3개월', '400%', ''],
        ['Phase 2', '₩15,000,000', '₩16,000,000 추가', '12개월', '107%', ''],
        ['Phase 3', '₩50,000,000', '₩60,000,000+ 추가', '12개월', '120%+', ''],
    ]

    for row_idx, row_data in enumerate(roi_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 5:
                cell.font = Font(bold=True, color=COLORS['success'])

    # 3년 TCO 비교
    ws.merge_cells('A11:F11')
    ws['A11'] = "3년 TCO 비교"
    apply_style(ws['A11'], create_section_title_style())

    tco_headers = ['시나리오', '초기 투자', '연간 운영비', '숨겨진 비용(3년)', '총 비용', '절감액']
    for col_idx, header in enumerate(tco_headers, 1):
        cell = ws.cell(row=12, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    tco_data = [
        ['현재 상태 유지', '₩0', '₩0', '₩330,000,000', '₩330,000,000', '-'],
        ['Phase 0-2 적용', '₩15,000,000', '₩9,000,000', '₩120,000,000', '₩144,000,000', '₩186,000,000'],
        ['Phase 0-3 적용', '₩50,000,000', '₩30,000,000', '₩60,000,000', '₩140,000,000', '₩190,000,000'],
    ]

    for row_idx, row_data in enumerate(tco_data, 13):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_data[0] == 'Phase 0-3 적용':
                cell.fill = PatternFill(start_color=COLORS['light_green'], fill_type='solid')
            if col_idx == 6 and value != '-':
                cell.font = Font(bold=True, color=COLORS['success'])

    return ws

def create_kpi_sheet(wb):
    """KPI 및 성공 지표 시트 생성"""
    ws = wb.create_sheet("KPI 성공지표")

    set_column_widths(ws, [5, 25, 20, 20, 20, 20])

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "성공 지표 (KPI)"
    apply_style(ws['A1'], create_title_style())
    ws.row_dimensions[1].height = 40

    # KPI 테이블
    ws.merge_cells('A3:F3')
    ws['A3'] = "Phase별 목표 지표"
    apply_style(ws['A3'], create_section_title_style())

    headers = ['Phase', '지표', 'AS-IS', 'TO-BE 목표', '측정 방법', '']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    kpi_data = [
        ['Phase 0', '긴급 대응 시간', '7일', '4시간', 'Fast Track 로그'],
        ['Phase 0', 'WISE 대기 시간', '2-3일', '1일 이내', 'SCR 시스템'],
        ['Phase 1', '검증 리드타임', '7일', '4.5일', 'SCR 시스템'],
        ['Phase 1', '코드 리뷰 적용률', '0%', '80%', 'PR 통계'],
        ['Phase 2', '빌드 자동화율', '0%', '100%', 'CI 시스템'],
        ['Phase 2', '테스트 커버리지', '0%', '30%', 'TcUnit 리포트'],
        ['Phase 3', '배포 자동화율', '0%', '80%', 'CD 시스템'],
        ['Phase 3', '평균 롤백 시간', '30-60분', '5-10분', '모니터링'],
    ]

    for row_idx, row_data in enumerate(kpi_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if col_idx == 3:
                cell.fill = PatternFill(start_color='FFC7CE', fill_type='solid')
            elif col_idx == 4:
                cell.fill = PatternFill(start_color='C6EFCE', fill_type='solid')
                cell.font = Font(bold=True)

    return ws

def main():
    """메인 함수"""
    print("엑셀 파일 생성 시작...")

    wb = Workbook()

    # 기본 시트 삭제 (나중에 생성)
    default_sheet = wb.active

    # 시트 생성 (순서대로)
    create_flowchart_sheet(wb)      # 1. 전체 흐름도 (첫 번째)
    create_workflow_sheet(wb)        # 2. 워크플로우 개요
    create_change_classification_sheet(wb)  # 3. 변경 분류 체계
    create_analysis_sheet(wb)        # 4. 장단점 분석
    create_as_is_to_be_sheet(wb)     # 5. AS-IS TO-BE
    create_roadmap_sheet(wb)         # 6. 실행 로드맵
    create_roi_sheet(wb)             # 7. ROI 분석
    create_kpi_sheet(wb)             # 8. KPI 성공지표

    # 기본 시트 삭제
    wb.remove(default_sheet)

    # 파일 저장
    output_path = r"D:\01. Vscode\Twincat\docs\TwinCAT_Workflow_Analysis.xlsx"
    wb.save(output_path)
    print(f"엑셀 파일 생성 완료: {output_path}")

    return output_path

if __name__ == "__main__":
    main()
